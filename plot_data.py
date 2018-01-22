import openpyxl
import matplotlib.pyplot as plt
import numpy

# open Excel spreadsheet
wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb.get_sheet_by_name('Global Flex')

# get number of last row
mx = sheet.max_row

flow = []

mx_row = sheet.cell(row=mx,column=4).value
while mx_row is None:
    mx = mx - 1
    mx_row = sheet.cell(row=mx,column=4).value
    
for i in range(2,mx+1):
    temp = sheet.cell(row=i,column=4).value
    flow.append(float(0 if temp is None else temp)/1e6)

num = len(flow)
maks = 0
mini = 0
som = 0
for i in range(0,num):
    som = som + flow[i]
    if flow[i] > maks:
        maks = flow[i]

    if flow[i] < mini:
        mini = flow[i]

avg = som/num

std = 0
for i in range(0,num):
    std = std + (flow[i] - avg)**2

std = (std/num)**(1/2)

print(avg)
print(std)
print(maks)
print(mini)

nom = 1000
maks = avg + 5*std
mini = avg - 5*std
db = (maks - mini)/nom
b_arr = []
f_arr = []
for i in range(1,nom+1):
    b = mini + i*db
    b_arr.append(b)
    count = 0
    for j in range(0,num):
        if (i == 1):
            if flow[j] < b:
                count = count + 1
        elif (i == nom+1):
            if (flow[j] > b):
                count = count + 1
        else:
            if (flow[j] < b and flow[j] > b-db):
                count = count + 1
    f_arr.append(count/num)

#plt.hist(flow,100,normed=1)
plt.plot(b_arr,f_arr)
plt.show()




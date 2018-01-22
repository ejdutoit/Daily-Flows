import openpyxl # for new .xlsx files
import xlrd # for old .xls files
import PyPDF2 # for PDF files
import csv

from openpyxl.styles import Alignment # cell alignment

import tkinter as tk # open dialog box to find file
from tkinter import filedialog

# to read from pdf
from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename() # get file name from dialog box

wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb.get_sheet_by_name('Global Flex')

#wread = xlrd.open_workbook('inflows  outflows(Manco).xls')
wread = xlrd.open_workbook(file_path)
sheet2 = wread.sheet_by_name('Fund totals')
#sheet2 = wread.sheet_by_index(0)

# read flows from fund totals
date = sheet2.cell(10,4).value # Python counting starts from 0
fund = sheet2.cell(101,6).value
flow = sheet2.cell(101,3).value

############################################################################################

# find page orientation and rotate if portrait
file_path = filedialog.askopenfilename() # get file name from dialog box
file1 = open(file_path,'rb')
read1 = PyPDF2.PdfFileReader(file1)
page1 = read1.getPage(0)
#page.rotateClockwise(90)
    
#pdf = PdfFileReader(file('Java Printing.pdf'))
page = read1.getPage(0).mediaBox
if page.getUpperRight_x() - page.getUpperLeft_x() > page.getUpperRight_y() - page.getLowerRight_y():
    print('Landscape')
else:
    print('Portrait')
    page1.rotateClockwise(90)
    
# save (rotated) page
file2 = PyPDF2.PdfFileWriter()
file2.addPage(page1)
result = open('Balances.pdf','wb')
file2.write(result)
result.close()
file1.close()

# read from pdf
fp = open('Balances.pdf', 'rb')
parser = PDFParser(fp)
doc = PDFDocument()
parser.set_document(doc)
doc.set_parser(parser)
doc.initialize('')
rsrcmgr = PDFResourceManager()
laparams = LAParams()
laparams.char_margin = 1.0
laparams.word_margin = 1.0
device = PDFPageAggregator(rsrcmgr, laparams=laparams)
interpreter = PDFPageInterpreter(rsrcmgr, device)
extracted_text = ''

for page in doc.get_pages():
    interpreter.process_page(page)
    layout = device.get_result()
    for lt_obj in layout:
        if isinstance(lt_obj, LTTextBox) or isinstance(lt_obj, LTTextLine):
            extracted_text += lt_obj.get_text()

# write to csv
outputFile = open('output.csv','w',newline='') # create csv file to write data to
outputFile.write(extracted_text)
outputFile.close()




# read data from csv file
inputFile = open('output.csv')
readFile = csv.reader(inputFile)
text = ''

for row in readFile:
    if readFile.line_num == 243: # Global Felx is line 243
        text = row

inputFile.close()
bank = ''
for i in range(0,len(text)):
    bank = bank + text[i]











############################################################################################

# write to data file
i = sheet.max_row

sheet.cell(row=i+1,column=1).value = date
sheet.cell(row=i+1,column=1).number_format = 'YYYY/MM/DD'
sheet.cell(row=i+1,column=1).alignment = Alignment(horizontal='left')
sheet.cell(row=i+1,column=2).value = fund
sheet.cell(row=i+1,column=2).style = 'Comma'
sheet.cell(row=i+1,column=3).value = float(bank)
sheet.cell(row=i+1,column=3).style = 'Comma'
sheet.cell(row=i+1,column=4).value = flow
sheet.cell(row=i+1,column=4).style = 'Comma'



wb.save('Data2.xlsx')
